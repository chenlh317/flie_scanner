"""
Scan a directory tree for large files and large subfolders, then export the
results to a timestamped Excel workbook with two formatted tables.

Run with `python scan_large_files.py` (edit the call at the bottom to point at
your directory and size cutoff).
"""

from __future__ import annotations

import os
import re
import subprocess
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# --- Constants ---------------------------------------------------------------

FILE_COLUMNS = ["file name", "file suffix", "sub folder dir", "size (bytes)", "size"]
FOLDER_COLUMNS = ["subfolder name", "sub folder dir", "size (bytes)", "size"]
SORT_COLUMN = "size (bytes)"

FILES_SHEET = "Large files"
FOLDERS_SHEET = "Large folders"
FILES_TABLE = "LargeFiles"
FOLDERS_TABLE = "LargeFolders"

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)

MAX_TAG_LEN = 80  # cap the length of the directory tag embedded in the filename


# --- Helpers -----------------------------------------------------------------

def human_size(num_bytes: float) -> str:
    """Return a human-readable size string (e.g. '1.23 GB')."""
    size = float(num_bytes)
    for unit in ("B", "KB", "MB", "GB", "TB", "PB"):
        if size < 1024 or unit == "PB":
            return f"{size:.2f} {unit}"
        size /= 1024
    return f"{size:.2f} PB"  # unreachable; keeps type checkers happy


def _relative_subfolder(folder: Path, root: Path) -> str:
    """Return `folder` relative to `root`; '' if they are the same path."""
    try:
        rel = str(folder.relative_to(root))
    except ValueError:
        return str(folder)
    return "" if rel == "." else rel


def _safe_tag(path: Path, max_len: int = MAX_TAG_LEN) -> str:
    """Build a filesystem-safe identifier derived from `path`."""
    tag = re.sub(r"[^A-Za-z0-9._-]+", "_", str(path)).strip("_")
    if len(tag) > max_len:
        tag = tag[-max_len:].lstrip("_")
    return tag or "root"


# --- Core scanning -----------------------------------------------------------

def scan(root: Path, cutoff_bytes: int) -> tuple[list[dict], list[dict]]:
    """Walk `root` and collect files and folders meeting the size cutoff.

    Folder sizes are recursive (they include every file beneath them).
    """
    file_rows: list[dict] = []
    folder_sizes: dict[Path, int] = defaultdict(int)

    for dirpath, _dirnames, filenames in os.walk(root):
        dir_path = Path(dirpath)
        for name in filenames:
            full = dir_path / name
            try:
                size = full.stat().st_size
            except OSError:
                continue

            # Add this file's size to every ancestor folder up to (and including) root.
            ancestor = dir_path
            while True:
                folder_sizes[ancestor] += size
                if ancestor == root:
                    break
                ancestor = ancestor.parent

            if size >= cutoff_bytes:
                file_rows.append({
                    "file name": name,
                    "file suffix": full.suffix.lower().lstrip("."),
                    "sub folder dir": _relative_subfolder(dir_path, root),
                    "size (bytes)": size,
                    "size": human_size(size),
                })

    folder_rows = [
        {
            "subfolder name": folder.name,
            "sub folder dir": _relative_subfolder(folder.parent, root),
            "size (bytes)": total,
            "size": human_size(total),
        }
        for folder, total in folder_sizes.items()
        if folder != root and total >= cutoff_bytes
    ]

    return file_rows, folder_rows


# --- Excel output ------------------------------------------------------------

def _autofit(worksheet: Worksheet) -> None:
    """Approximate Excel's autofit by sizing each column to its widest cell."""
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_len = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def _add_table(worksheet: Worksheet, table_name: str, n_rows: int, n_cols: int) -> None:
    """Wrap the written range in a styled Excel Table (ListObject)."""
    if n_cols == 0:
        return
    # Excel tables require at least one data row; pad an empty row if needed.
    last_row = max(n_rows + 1, 2)
    ref = f"A1:{get_column_letter(n_cols)}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TABLE_STYLE
    worksheet.add_table(table)


def _to_dataframe(rows: Iterable[dict], columns: list[str]) -> pd.DataFrame:
    return (
        pd.DataFrame(rows, columns=columns)
        .sort_values(SORT_COLUMN, ascending=False)
        .reset_index(drop=True)
    )


def _write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet: str, table: str) -> None:
    df.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    _add_table(ws, table, len(df), len(df.columns))
    _autofit(ws)


def _open_folder(path: Path) -> None:
    """Open the containing folder of `path` in the OS file explorer."""
    folder = path.parent
    try:
        if sys.platform.startswith("win"):
            os.startfile(folder)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(folder)])
        else:
            subprocess.Popen(["xdg-open", str(folder)])
    except OSError as exc:
        print(f"Could not open folder {folder}: {exc}", file=sys.stderr)


# --- Entry point -------------------------------------------------------------

def main(directory: str, cutoff_mb: float = 500.0) -> Path:
    root = Path(directory).expanduser().resolve()
    if not root.is_dir():
        print(f"Error: {root} is not a directory", file=sys.stderr)
        sys.exit(1)

    cutoff_bytes = int(cutoff_mb * 1024 * 1024)
    script_dir = Path(__file__).resolve().parent
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = script_dir / f"large_files_{_safe_tag(root)}_{timestamp}.xlsx"

    print(f"Scanning {root} for files/folders >= {cutoff_mb} MB ...")
    file_rows, folder_rows = scan(root, cutoff_bytes)
    print(f"Found {len(file_rows)} large file(s) and {len(folder_rows)} large folder(s).")

    files_df = _to_dataframe(file_rows, FILE_COLUMNS)
    folders_df = _to_dataframe(folder_rows, FOLDER_COLUMNS)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_sheet(writer, files_df, FILES_SHEET, FILES_TABLE)
        _write_sheet(writer, folders_df, FOLDERS_SHEET, FOLDERS_TABLE)

    print(f"Saved: {output_path}")
    _open_folder(output_path)
    return output_path


if __name__ == "__main__":
    main(
        directory=r"E:",
        cutoff_mb=500.0,
    )
